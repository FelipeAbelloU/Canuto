"""Evalua el modelo fine-tuneado sobre el conjunto de prueba (data/dataset/test.json)
y registra los resultados en una bitacora Excel (una fila por corrida).

Metricas (todas contra la respuesta de referencia del test):
  - BERTScore-F1 y similitud coseno: parecido semantico.
  - ROUGE-L / BLEU / CIDEr: solapamiento de texto.
  - Perplejidad / perdida: que tan bien el modelo predice la respuesta esperada.
  - Entropia: confianza del modelo (entropia media de su distribucion).
  - Informacion mutua: divergencia de vocabulario pred vs ref (informativa).
  - Exactitud de citas: si menciona el documento correcto (No. X de AAAA).
  - Tasa de alucinacion: sobre preguntas fuera de dominio, cuantas veces inventa.

La generacion es greedy (temperatura 0) para que las corridas sean comparables.
El modelo se toma de config.yaml (checkpoint_path) o de --checkpoint.

Uso:
    python scripts/evaluate.py --checkpoint data/checkpoints/7b-e3-lr2e4-r16 --epochs 3 --lr 2e-4 --limit 200
"""
import argparse
import json
import math
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.factory import create_pipeline

# Frases con las que el modelo reconoce que no tiene la informacion (no alucina)
_FRASES_NO_SE = [
    "no tengo", "no dispongo", "no cuento con", "no encuentro", "no aparece",
    "no está en", "no se encuentra", "no hay información", "no tengo información",
    "no tengo certeza", "no forma parte", "no puedo", "no está disponible",
    "no corresponde", "fuera de mi", "no dispongo de",
]

# Cita de un documento: "No. 10 de 2015", "N° 007 de 1990", "10 de 2015"
_CITA_RE = re.compile(r"N[°ºo\.]*\s*0*(\d{1,4})\s+de\s+(\d{4})", re.IGNORECASE)


def parse_args():
    p = argparse.ArgumentParser(description="Evaluacion del modelo + bitacora")
    p.add_argument("--test", default=str(ROOT / "data/dataset/test.json"))
    p.add_argument("--fuera-dominio", default=str(ROOT / "data/dataset/preguntas_fuera_dominio.json"))
    p.add_argument("--config", default=str(ROOT / "config/config.yaml"))
    p.add_argument("--out", default=str(ROOT / "data/dataset/resultados_evaluacion.json"))
    p.add_argument("--bitacora", default=str(ROOT / "data/bitacora.xlsx"),
                   help="Excel donde se agrega una fila por corrida")
    p.add_argument("--limit", type=int, default=0, help="Evaluar solo N ejemplos (0 = todos)")
    p.add_argument("--checkpoint", default=None,
                   help="Adaptador a evaluar (ignora el checkpoint_path de config.yaml)")
    p.add_argument("--epochs", default=None, help="Epocas de entrenamiento (para la bitacora)")
    p.add_argument("--lr", default=None, help="Learning rate de entrenamiento (para la bitacora)")
    return p.parse_args()


# --------------------------------------------------------------------------- #
# Metricas de texto (no necesitan el modelo)
# --------------------------------------------------------------------------- #
def citas(texto):
    """Conjunto de citas (numero, anio) encontradas en un texto."""
    return {(n.lstrip("0") or "0", a) for n, a in _CITA_RE.findall(texto)}


def reconoce_no_saber(texto):
    t = texto.lower()
    return any(f in t for f in _FRASES_NO_SE)


def _ngramas(tokens, n):
    return [tuple(tokens[i:i + n]) for i in range(len(tokens) - n + 1)]


def cider(preds, refs, n_max=4):
    """CIDEr simplificado: coseno de vectores de n-gramas ponderados por TF-IDF,
    promediado para n=1..4. El IDF sale del corpus de referencias. Con una sola
    referencia por ejemplo es en esencia un coseno TF-IDF de n-gramas (informativo)."""
    refs_tok = [r.lower().split() for r in refs]
    preds_tok = [p.lower().split() for p in preds]
    N = len(refs_tok)
    if N == 0:
        return 0.0
    puntajes = []
    for n in range(1, n_max + 1):
        refs_ng = [Counter(_ngramas(rt, n)) for rt in refs_tok]
        df = Counter()
        for ng in refs_ng:
            for g in ng:
                df[g] += 1

        def idf(g):
            return math.log((N + 1) / (df.get(g, 0) + 1)) + 1

        def vec(counter):
            return {g: c * idf(g) for g, c in counter.items()}

        def cos(a, b):
            if not a or not b:
                return 0.0
            dot = sum(v * b.get(g, 0) for g, v in a.items())
            na = math.sqrt(sum(v * v for v in a.values()))
            nb = math.sqrt(sum(v * v for v in b.values()))
            return dot / (na * nb) if na and nb else 0.0

        s = sum(cos(vec(Counter(_ngramas(pt, n))), vec(rng))
                for pt, rng in zip(preds_tok, refs_ng))
        puntajes.append(s / N)
    return sum(puntajes) / len(puntajes)


def info_mutua(pred, ref):
    """MI entre 'de que texto (pred/ref)' y 'que token'. Es una DIVERGENCIA de
    vocabulario: 0 cuando ambos usan las palabras igual, sube cuando difieren
    (mayor = mas distintos). Informativa, NO es criterio de seleccion."""
    tp, tr = pred.lower().split(), ref.lower().split()
    if not tp or not tr:
        return 0.0
    cp, cr = Counter(tp), Counter(tr)
    N = len(tp) + len(tr)
    mi = 0.0
    for t in set(tp) | set(tr):
        pt = (cp[t] + cr[t]) / N
        for cnt, tot in ((cp[t], len(tp)), (cr[t], len(tr))):
            if cnt:
                pst = cnt / N
                ps = tot / N
                mi += pst * math.log(pst / (ps * pt))
    return mi


def coseno_semantico(preds, refs):
    """Similitud coseno media entre embeddings de la respuesta y la referencia."""
    from sentence_transformers import SentenceTransformer
    modelo = SentenceTransformer("sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2")
    ep = modelo.encode(preds, convert_to_tensor=True, normalize_embeddings=True)
    er = modelo.encode(refs, convert_to_tensor=True, normalize_embeddings=True)
    return float((ep * er).sum(dim=1).mean())


# --------------------------------------------------------------------------- #
# Metricas que necesitan el modelo (forward pass)
# --------------------------------------------------------------------------- #
def perplejidad_entropia(model, tokenizer, system_prompt, pregunta, respuesta, device, max_len=1024):
    """Un forward pass sobre (prompt + respuesta de referencia), en el MISMO formato
    de chat con que se entreno. Devuelve (perdida, entropia) sobre los tokens de la
    respuesta. La perplejidad = exp(media de perdida) se calcula fuera (mas estable).
    Se trunca a max_len (el seq_len del entrenamiento) para no medir fuera de distribucion
    ni reventar la VRAM con respuestas muy largas (Actas)."""
    import torch
    base = [{"role": "system", "content": system_prompt},
            {"role": "user", "content": pregunta}]
    full = base + [{"role": "assistant", "content": respuesta}]
    prompt_text = tokenizer.apply_chat_template(base, tokenize=False, add_generation_prompt=True)
    full_text = tokenizer.apply_chat_template(full, tokenize=False, add_generation_prompt=False)

    # add_special_tokens=False: el template ya trae los tokens especiales como texto,
    # asi el prompt tokeniza igual como prefijo del texto completo (alineacion exacta).
    p = tokenizer(prompt_text, add_special_tokens=False, return_tensors="pt").input_ids.shape[1]
    target = "cpu" if device == "cpu" else "cuda"
    full_ids = tokenizer(full_text, add_special_tokens=False, return_tensors="pt").input_ids[:, :max_len].to(target)

    with torch.no_grad():
        logits = model(full_ids).logits[0, :-1, :]   # predice el token t desde t-1
    labels = full_ids[0, 1:]
    ini = max(p - 1, 0)                                # solo la respuesta
    sel_logits, sel_labels = logits[ini:], labels[ini:]
    if sel_labels.numel() == 0:
        return 0.0, 0.0
    logprobs = torch.log_softmax(sel_logits.float(), dim=-1)
    nll = -logprobs[range(sel_labels.shape[0]), sel_labels].mean().item()
    entropia = -(logprobs.exp() * logprobs).sum(dim=-1).mean().item()
    return nll, entropia


def hiperparams_adapter(checkpoint_dir):
    """Lee rank / alpha / dropout del adapter_config.json (si es un adaptador LoRA)."""
    cfg = Path(checkpoint_dir) / "adapter_config.json"
    if cfg.exists():
        d = json.loads(cfg.read_text(encoding="utf-8"))
        return d.get("r"), d.get("lora_alpha"), d.get("lora_dropout")
    return None, None, None


# --------------------------------------------------------------------------- #
# Bitacora
# --------------------------------------------------------------------------- #
COLUMNAS = [
    "fecha", "checkpoint", "n_prueba", "epochs", "learning_rate",
    "rank", "lora_alpha", "lora_dropout",
    "bertscore_f1", "coseno_sem", "rouge_l", "bleu", "cider",
    "perdida", "perplejidad", "entropia", "info_mutua",
    "exactitud_citas", "tasa_alucinacion",
]


def escribir_bitacora(path, fila):
    """Agrega una fila al Excel (crea el archivo con encabezados la primera vez)."""
    from openpyxl import Workbook, load_workbook
    p = Path(path)
    if p.exists():
        wb = load_workbook(p)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "experimentos"
        ws.append(COLUMNAS)
    ws.append([fila.get(c) for c in COLUMNAS])
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)


def generar_respuestas(pipeline, preguntas):
    """Pasa cada pregunta por el modelo (sin memoria entre preguntas)."""
    respuestas = []
    for i, q in enumerate(preguntas, 1):
        pipeline.reset_history()
        respuestas.append(pipeline.query(q)["answer"])
        print(f"  {i}/{len(preguntas)}", end="\r")
    print()
    return respuestas


def main():
    args = parse_args()

    pipeline = create_pipeline(args.config, checkpoint_override=args.checkpoint)
    if pipeline.model is None or not pipeline.model.is_available():
        print("No hay checkpoint configurado en config.yaml (model.checkpoint_path).")
        sys.exit(1)
    pipeline.model.temperature = 0   # greedy: corridas comparables
    print(f"Modelo: {pipeline.model.checkpoint_path} (device: {pipeline.model.device})")

    with open(args.test, encoding="utf-8") as f:
        data = json.load(f)
    if args.limit:
        data = data[:args.limit]

    preguntas = [r["instruction"] for r in data]
    esperadas = [r["output"] for r in data]

    print(f"Generando respuestas para {len(preguntas)} preguntas de prueba...")
    predichas = generar_respuestas(pipeline, preguntas)

    # --- BERTScore (parecido semantico) ---
    from bert_score import score as bert_score
    _, _, f1 = bert_score(predichas, esperadas, lang="es", rescale_with_baseline=False)
    bertscore_f1 = float(f1.mean())

    # --- Coseno semantico ---
    coseno = coseno_semantico(predichas, esperadas)

    # --- ROUGE-L ---
    from rouge_score import rouge_scorer
    scorer = rouge_scorer.RougeScorer(["rougeL"], use_stemmer=False)
    rouge_l = sum(scorer.score(ref, pred)["rougeL"].fmeasure
                  for ref, pred in zip(esperadas, predichas)) / len(data)

    # --- BLEU ---
    import sacrebleu
    bleu = sacrebleu.corpus_bleu(predichas, [esperadas]).score

    # --- CIDEr ---
    cider_score = cider(predichas, esperadas)

    # --- Informacion mutua (divergencia de vocabulario) ---
    im = sum(info_mutua(pred, ref) for pred, ref in zip(predichas, esperadas)) / len(data)

    # --- Perplejidad / perdida / entropia (forward pass) ---
    print("Calculando perplejidad y entropia...")
    hf_model = pipeline.model.hf_model
    tok = pipeline.model.tokenizer
    dev = pipeline.model.device
    nlls, entrs = [], []
    for i, (q, ref) in enumerate(zip(preguntas, esperadas), 1):
        nll, ent = perplejidad_entropia(hf_model, tok, pipeline.system_prompt, q, ref, dev)
        nlls.append(nll); entrs.append(ent)
        print(f"  {i}/{len(preguntas)}", end="\r")
    print()
    perdida = sum(nlls) / len(nlls)
    perplejidad = math.exp(perdida)              # exp(media) es mas estable que media(exp)
    entropia = sum(entrs) / len(entrs)

    # --- Exactitud de citas ---
    con_cita = aciertos = 0
    for ref, pred in zip(esperadas, predichas):
        objetivo = citas(ref)
        if objetivo:
            con_cita += 1
            if objetivo & citas(pred):
                aciertos += 1
    exactitud_citas = (aciertos / con_cita) if con_cita else 0.0

    # --- Tasa de alucinacion (preguntas fuera de dominio) ---
    with open(args.fuera_dominio, encoding="utf-8") as f:
        fuera = json.load(f)
    print(f"Generando respuestas para {len(fuera)} preguntas fuera de dominio...")
    resp_fuera = generar_respuestas(pipeline, fuera)
    alucina = sum(0 if reconoce_no_saber(r) else 1 for r in resp_fuera)
    tasa_alucinacion = alucina / len(fuera)

    # --- Hiperparametros para la bitacora ---
    ckpt_dir = args.checkpoint or pipeline.model.checkpoint_path
    rank, alpha, dropout = hiperparams_adapter(ckpt_dir)

    resultados = {
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "checkpoint": pipeline.model.name,
        "n_prueba": len(data),
        "epochs": args.epochs,
        "learning_rate": args.lr,
        "rank": rank,
        "lora_alpha": alpha,
        "lora_dropout": dropout,
        "bertscore_f1": round(bertscore_f1, 4),
        "coseno_sem": round(coseno, 4),
        "rouge_l": round(rouge_l, 4),
        "bleu": round(bleu, 2),
        "cider": round(cider_score, 4),
        "perdida": round(perdida, 4),
        "perplejidad": round(perplejidad, 2),
        "entropia": round(entropia, 4),
        "info_mutua": round(im, 4),
        "exactitud_citas": round(exactitud_citas, 4),
        "tasa_alucinacion": round(tasa_alucinacion, 4),
    }

    print("\n=== Resultados ===")
    for k, v in resultados.items():
        print(f"  {k}: {v}")

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)
    escribir_bitacora(args.bitacora, resultados)
    print(f"\nGuardado en: {args.out}")
    print(f"Fila agregada a la bitacora: {args.bitacora}")


if __name__ == "__main__":
    main()
